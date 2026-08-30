import datetime
import os
import re
import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

URL_CONSULTA = "TU_URL_AQUI"  # Reemplazar por la URL del portal municipal
EXCEL_FILE = "propiedades.xlsx"
TXT_OUTPUT = "reporte_deudas_sanitarios.txt"


def parse_monto(texto_monto):
    """Limpia cadenas de texto con montos y devuelve un float."""
    if not texto_monto:
        return 0.0
    s = re.sub(r'[^\d.,]', '', str(texto_monto))
    if not s:
        return 0.0
    if ',' in s and '.' in s:
        if s.rfind('.') > s.rfind(','):
            s = s.replace(',', '')
        else:
            s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        parts = s.split(',')
        if len(parts[-1]) == 2:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def extraer_registros_excel(archivo):
    """Lee el Excel, filtra títulos, completa propietarios heredados y extrae cuentas."""
    wb = openpyxl.load_workbook(archivo, data_only=True)
    sheet = wb.active

    registros = []
    propietario_actual = ""

    for r in range(4, sheet.max_row + 1):
        u = sheet.cell(row=r, column=1).value
        d = sheet.cell(row=r, column=2).value
        c1 = sheet.cell(row=r, column=3).value
        c2 = sheet.cell(row=r, column=4).value

        # Actualizar propietario si existe en la fila
        if u and str(u).strip():
            u_clean = str(u).strip().replace('\n', ' ')
            if "Propietario" not in u_clean and "Estado de Pago" not in u_clean:
                propietario_actual = u_clean

        dir_str = str(d).strip() if d else ""

        # Ignorar filas de títulos
        if "Dirección" in dir_str or "Ubicación" in dir_str:
            continue

        for c_raw in [c1, c2]:
            if not c_raw:
                continue
            c_str = str(c_raw).strip()
            if "$ 0,00" in c_str or c_str.lower() == "none" or not c_str:
                continue

            # Buscar secuencias de 5 a 10 dígitos (cuentas válidas)
            cuentas_encontradas = re.findall(r'\b\d{5,10}\b', c_str)
            for acct in cuentas_encontradas:
                registros.append({
                    "propietario": propietario_actual,
                    "direccion": dir_str,
                    "cuenta": acct
                })

    return registros


def iniciar_driver():
    """Inicializa una nueva instancia de ChromeDriver."""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def procesar_consultas():
    registros = extraer_registros_excel(EXCEL_FILE)
    print(f"Cuentas cargadas para procesar: {len(registros)}")

    driver = iniciar_driver()
    lineas_reporte = []

    fecha_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    lineas_reporte.append("=== REPORTE DE DEUDAS - SERVICIOS SANITARIOS ===")
    lineas_reporte.append(f"Fecha de consulta: {fecha_hora}")
    lineas_reporte.append("=" * 55 + "\n")

    for idx, reg in enumerate(registros, 1):
        user = reg["propietario"]
        direccion = reg["direccion"]
        cuenta = reg["cuenta"]

        print(f"[{idx}/{len(registros)}] Procesando: {user} | Cuenta: {cuenta}")

        intento = 0
        exito = False
        comprobantes = []
        error_msg = None

        while intento < 2 and not exito:
            intento += 1
            try:
                driver.get(URL_CONSULTA)

                # 1. Localizar input de cuenta y botón de búsqueda (Ajustar Selectores)
                input_cuenta = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "P1_CUENTA"))
                )
                input_cuenta.clear()
                input_cuenta.send_keys(cuenta)

                btn_buscar = driver.find_element(By.ID, "P1_BUSCAR")
                btn_buscar.click()

                # 2. Esperar tabla de resultados o mensaje de sin deudas
                time.sleep(2)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "a-IRR-table"))
                )

                # 3. Leer filas de la grilla
                filas = driver.find_elements(By.XPATH, "//table[contains(@class,'a-IRR-table')]/tbody/tr")
                for f in filas:
                    cols = f.find_elements(By.TAG_NAME, "td")
                    if len(cols) >= 5:
                        concepto = cols[1].text.strip()
                        vencimiento = cols[3].text.strip()
                        monto_txt = cols[4].text.strip()

                        if "SERV" in concepto.upper() or "SANITAR" in concepto.upper():
                            monto_val = parse_monto(monto_txt)
                            comprobantes.append({
                                "concepto": concepto,
                                "vencimiento": vencimiento,
                                "monto_txt": monto_txt,
                                "monto_val": monto_val
                            })

                exito = True

            except Exception as e:
                error_msg = str(e).split('\n')[0]
                if "invalid session id" in error_msg or "disconnected" in error_msg:
                    print("Sesión de navegador perdida. Reiniciando driver...")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = iniciar_driver()

        # Armar texto del reporte para la cuenta
        total_deuda = sum(c["monto_val"] for c in comprobantes)

        lineas_reporte.append(f"Usuario: {user}")
        lineas_reporte.append(f"Dirección: {direccion}")
        lineas_reporte.append(f"N° Cuenta: {cuenta}")
        lineas_reporte.append(f"Total Deuda SERV.SANITAR: ${total_deuda:,.2f}")

        if comprobantes:
            for c in comprobantes:
                lineas_reporte.append(
                    f"  • {c['concepto']} (Venc: {c['vencimiento']}): Total = {c['monto_txt']}"
                )
        elif error_msg:
            lineas_reporte.append(f"  • Error en lectura: {error_msg}")
        else:
            lineas_reporte.append("  • Sin deudas registradas.")

        lineas_reporte.append("-" * 55)

    try:
        driver.quit()
    except Exception:
        pass

    # Guardar archivo .txt
    with open(TXT_OUTPUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas_reporte))

    print(f"\nProceso finalizado. Reporte generado en: {TXT_OUTPUT}")


if __name__ == "__main__":
    procesar_consultas()

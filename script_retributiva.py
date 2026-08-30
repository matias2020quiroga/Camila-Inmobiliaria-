import time
import openpyxl
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

EXCEL_FILE = "propiedades.xlsx"
TXT_OUTPUT = "resultados_retributiva_servicios.txt"
URL_TANDIL = "https://www.autogestion.tandil.gov.ar/apex/f?p=114:101"

def formatear_texto(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str

def parse_monto(texto_monto):
    """Convierte importes en texto ($85,535.92 o $85.535,92) a float correctamente."""
    if not texto_monto:
        return 0.0
    s = re.sub(r'[^\d.,]', '', str(texto_monto))
    if not s:
        return 0.0
    if ',' in s and '.' in s:
        if s.rfind('.') > s.rfind(','):
            s = s.replace(',', '')
        else:
            s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        parts = s.split(',')
        if len(parts[-1]) == 2:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return 0.0

def consultar_retributiva_servicios():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb.active
    except Exception as e:
        print(f"Error al abrir Excel: {e}")
        return

    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(5)

    with open(TXT_OUTPUT, "w", encoding="utf-8") as f_out:
        f_out.write("=== REPORTE DE DEUDAS - RETRIBUTIVA DE SERVICIOS ===\n")
        f_out.write(f"Fecha de consulta: {time.strftime('%d/%m/%Y %H:%M')}\n")
        f_out.write("=" * 55 + "\n\n")

        try:
            for row in range(2, sheet.max_row + 1):
                val_usuario = sheet.cell(row=row, column=1).value
                val_direccion = sheet.cell(row=row, column=2).value
                # CAMBIO: Toma la cuenta de la Columna 4
                val_servicio = sheet.cell(row=row, column=4).value

                num_cuenta = formatear_texto(val_servicio)

                if not num_cuenta or num_cuenta.lower() in ["none", "servicio", "cuenta", ""]:
                    continue

                usuario = formatear_texto(val_usuario) or f"Usuario_{row}"
                direccion = formatear_texto(val_direccion) or "Sin Dirección"

                # 1. Cargar la página inicial
                driver.get(URL_TANDIL)
                time.sleep(1.5)

                # 2. Completar Nro. Cuenta
                try:
                    input_cuenta = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text']"))
                    )
                    input_cuenta.clear()
                    input_cuenta.send_keys(num_cuenta)
                except Exception:
                    continue

                # 3. Seleccionar 'Retributiva de Servicios'
                try:
                    dropdown_elem = driver.find_element(By.TAG_NAME, "select")
                    select_tasa = Select(dropdown_elem)
                    for option in select_tasa.options:
                        if "retributiva" in option.text.lower():
                            select_tasa.select_by_visible_text(option.text)
                            break
                except Exception:
                    pass

                time.sleep(1)

                # 4. Iniciar Sesión
                try:
                    btn_iniciar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Iniciar') or contains(., 'Sesión')] | //input[@type='submit' or @type='button']"))
                    )
                    btn_iniciar.click()
                except Exception:
                    driver.execute_script("document.querySelector('button').click();")

                time.sleep(3)

                # 5. FORZAR CLIC EN 'GENERAR COMPR. PAGO'
                try:
                    driver.execute_script("""
                        var elem = document.querySelectorAll('a, span, li');
                        for (var i = 0; i < elem.length; i++) {
                            if (elem[i].innerText.includes('Generar compr')) {
                                elem[i].click();
                                break;
                            }
                        }
                    """)
                    time.sleep(3)
                except Exception:
                    pass

                # 6. Leer la tabla de comprobantes de pago
                registros_deuda = []

                try:
                    WebDriverWait(driver, 12).until(
                        EC.presence_of_element_located((By.XPATH, "//table//td[contains(., 'RETRIB') or contains(., '$')] | //td"))
                    )
                    time.sleep(1.5)

                    filas_tabla = driver.find_elements(By.XPATH, "//tr")
                    for fila in filas_tabla:
                        texto_fila = fila.text
                        if "RETRIB" in texto_fila.upper() or "$" in texto_fila:
                            columnas = fila.find_elements(By.TAG_NAME, "td")
                            
                            # Estructura: Checkbox(0), Cuenta(1), Vencimiento(2), Tasa(3), Total(4)
                            if len(columnas) >= 5:
                                vencimiento = columnas[2].text.strip()
                                tasa_info = columnas[3].text.strip()
                                importe_raw = columnas[4].text.strip()
                                monto_float = parse_monto(importe_raw)

                                registros_deuda.append({
                                    "texto": f"  • {tasa_info} (Venc: {vencimiento}): Total = ${importe_raw}",
                                    "monto": monto_float
                                })
                except Exception as ex_tabla:
                    registros_deuda.append({"texto": f"  • Error en lectura de tabla: {ex_tabla}", "monto": 0.0})

                # --- SELECCIONAR ÚNICAMENTE LAS 2 DEUDAS MÁS RECIENTES (LAS DE MÁS ARRIBA) ---
                ultimas_deudas = registros_deuda[:2]
                monto_total_retributiva = sum(item["monto"] for item in ultimas_deudas)
                detalles_deuda = [item["texto"] for item in ultimas_deudas]

                # 7. Escribir resultados
                if detalles_deuda:
                    resumen = "\n".join(detalles_deuda)
                    linea_monto = f"Total Deuda RETRIBUTIVA DE SERVICIOS (Últimos 2): ${monto_total_retributiva:,.2f}\n{resumen}"
                else:
                    linea_monto = "Sin Deuda en RETRIBUTIVA DE SERVICIOS / No Encontrado"

                registro = (
                    f"Usuario: {usuario}\n"
                    f"Dirección: {direccion}\n"
                    f"N° Cuenta: {num_cuenta}\n"
                    f"{linea_monto}\n"
                    f"-------------------------------------------------------\n"
                )
                f_out.write(registro)
                f_out.flush()

        finally:
            driver.quit()

if __name__ == "__main__":
    consultar_retributiva_servicios()
